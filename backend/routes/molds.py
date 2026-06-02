from flask import Blueprint, request, jsonify
from models import get_db_connection
from utils import login_required, validate_mold_no, validate_mold_name, validate_positive_int
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import traceback

molds_bp = Blueprint('molds', __name__)

STATUS_SQL = """
    SELECT m.*, ms.current_count, ms.last_maintain_count, ms.last_maintain_date,
           ms.total_production, (ms.total_production + ms.current_count) as actual_total_production,
           (ms.current_count - ms.last_maintain_count) as used_count,
           CASE
               WHEN (ms.current_count - ms.last_maintain_count) > m.maintain_interval THEN 'needs_maintenance'
               WHEN (ms.current_count - ms.last_maintain_count) >= m.maintain_interval * 0.9 THEN 'warning'
               ELSE 'normal'
           END as status
    FROM molds m LEFT JOIN mold_status ms ON m.id = ms.mold_id
"""

def _fmt_row(row):
    if row:
        for k in ['last_maintain_date', 'reset_date', 'maintain_date', 'created_at', 'updated_at']:
            if k in row and row[k]:
                row[k] = str(row[k])
    return row

def _fmt_rows(rows):
    return [_fmt_row(dict(r)) for r in rows]

# ─── Mold CRUD ───
@molds_bp.route('/api/molds', methods=['GET'])
@login_required
def get_molds():
    search = request.args.get('search', '').strip()
    page = max(1, int(request.args.get('page', 1)))
    page_size = max(1, min(100, int(request.args.get('pageSize', 10))))
    offset = (page - 1) * page_size

    conn = get_db_connection()
    try:
        c = conn.cursor()
        where, params = "", []
        if search:
            where = "WHERE m.mold_no LIKE %s OR m.mold_name LIKE %s"
            params = [f"%{search}%", f"%{search}%"]

        c.execute(f"SELECT COUNT(*) as total FROM molds m {where}", params)
        total = c.fetchone()['total']

        sql = f"""SELECT m.*, ms.current_count, ms.last_maintain_count, ms.last_maintain_date,
                   ms.total_production, (ms.total_production + ms.current_count) as actual_total_production
            FROM molds m LEFT JOIN mold_status ms ON m.id = ms.mold_id
            {where} ORDER BY m.created_at DESC LIMIT %s OFFSET %s"""
        c.execute(sql, params + [page_size, offset])
        rows = c.fetchall()
        return jsonify({'code': 200, 'data': {'list': _fmt_rows(rows), 'total': total, 'page': page, 'pageSize': page_size}})
    finally: conn.close()

@molds_bp.route('/api/molds', methods=['POST'])
@login_required
def create_mold():
    data = request.get_json() or {}
    mold_no = data.get('mold_no', '').strip().upper()
    mold_name = data.get('mold_name', '').strip()
    maintain_interval = data.get('maintain_interval')

    errors = {}
    err = validate_mold_no(mold_no)
    if err: errors['mold_no'] = err
    err = validate_mold_name(mold_name)
    if err: errors['mold_name'] = err
    err = validate_positive_int(maintain_interval, '保养模次间隔')
    if err: errors['maintain_interval'] = err
    if errors: return jsonify({'code': 400, 'message': '输入有误', 'errors': errors}), 400

    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT id FROM molds WHERE mold_no = %s", (mold_no,))
        if c.fetchone(): return jsonify({'code': 400, 'message': '模具编号已存在', 'errors': {'mold_no': '模具编号已存在'}}), 400

        c.execute("INSERT INTO molds (mold_no, mold_name, maintain_interval) VALUES (%s, %s, %s)", (mold_no, mold_name, int(maintain_interval)))
        new_id = c.lastrowid
        c.execute("INSERT INTO mold_status (mold_id, current_count, last_maintain_count, total_production) VALUES (%s, 0, 0, 0)", (new_id,))
        conn.commit()

        c.execute("SELECT m.*, ms.current_count, ms.last_maintain_count, ms.last_maintain_date, ms.total_production FROM molds m LEFT JOIN mold_status ms ON m.id = ms.mold_id WHERE m.id = %s", (new_id,))
        return jsonify({'code': 200, 'message': '添加成功', 'data': _fmt_row(c.fetchone())})
    except Exception as e: conn.rollback(); return jsonify({'code': 500, 'message': str(e)}), 500
    finally: conn.close()

@molds_bp.route('/api/molds/<int:mold_id>', methods=['PUT'])
@login_required
def update_mold(mold_id):
    data = request.get_json() or {}
    mold_no = data.get('mold_no', '').strip().upper()
    mold_name = data.get('mold_name', '').strip()
    maintain_interval = data.get('maintain_interval')

    errors = {}
    err = validate_mold_no(mold_no)
    if err: errors['mold_no'] = err
    err = validate_mold_name(mold_name)
    if err: errors['mold_name'] = err
    err = validate_positive_int(maintain_interval, '保养模次间隔')
    if err: errors['maintain_interval'] = err
    if errors: return jsonify({'code': 400, 'message': '输入有误', 'errors': errors}), 400

    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT id FROM molds WHERE id = %s", (mold_id,))
        if not c.fetchone(): return jsonify({'code': 404, 'message': '模具不存在'}), 404
        c.execute("SELECT id FROM molds WHERE mold_no = %s AND id != %s", (mold_no, mold_id))
        if c.fetchone(): return jsonify({'code': 400, 'message': '模具编号已存在', 'errors': {'mold_no': '模具编号已存在'}}), 400

        c.execute("UPDATE molds SET mold_no = %s, mold_name = %s, maintain_interval = %s WHERE id = %s",
                  (mold_no, mold_name, int(maintain_interval), mold_id))
        conn.commit()
        c.execute("""SELECT m.*, ms.current_count, ms.last_maintain_count, ms.last_maintain_date, ms.total_production
            FROM molds m LEFT JOIN mold_status ms ON m.id = ms.mold_id WHERE m.id = %s""", (mold_id,))
        return jsonify({'code': 200, 'message': '更新成功', 'data': _fmt_row(c.fetchone())})
    except Exception as e: conn.rollback(); return jsonify({'code': 500, 'message': str(e)}), 500
    finally: conn.close()

@molds_bp.route('/api/molds/<int:mold_id>', methods=['DELETE'])
@login_required
def delete_mold(mold_id):
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("DELETE FROM molds WHERE id = %s", (mold_id,))
        conn.commit()
        return jsonify({'code': 200, 'message': '删除成功'})
    except Exception as e: conn.rollback(); return jsonify({'code': 500, 'message': str(e)}), 500
    finally: conn.close()

# ─── Focus ───
@molds_bp.route('/api/molds/<int:mold_id>/focus', methods=['PUT'])
@login_required
def set_focus(mold_id):
    data = request.get_json() or {}
    is_focused = 1 if data.get('is_focused') else 0
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("UPDATE molds SET is_focused = %s WHERE id = %s", (is_focused, mold_id))
        conn.commit()
        return jsonify({'code': 200, 'message': '设置成功'})
    except Exception as e: conn.rollback(); return jsonify({'code': 500, 'message': str(e)}), 500
    finally: conn.close()

@molds_bp.route('/api/molds/focused', methods=['GET'])
@login_required
def get_focused_molds():
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute(f"{STATUS_SQL} WHERE m.is_focused = 1 ORDER BY m.mold_no")
        return jsonify({'code': 200, 'data': _fmt_rows(c.fetchall())})
    finally: conn.close()

# ─── Status ───
@molds_bp.route('/api/molds/status', methods=['GET'])
@login_required
def get_mold_status():
    search = request.args.get('search', '').strip()
    page = max(1, int(request.args.get('page', 1)))
    page_size = max(1, min(100, int(request.args.get('pageSize', 10))))
    offset = (page - 1) * page_size

    conn = get_db_connection()
    try:
        c = conn.cursor()
        where, params = "", []
        if search:
            where = "WHERE m.mold_no LIKE %s OR m.mold_name LIKE %s"
            params = [f"%{search}%", f"%{search}%"]

        c.execute(f"SELECT COUNT(*) as total FROM molds m LEFT JOIN mold_status ms ON m.id = ms.mold_id {where}", params)
        total = c.fetchone()['total']

        sql = f"{STATUS_SQL} {where} ORDER BY FIELD(status, 'needs_maintenance', 'warning', 'normal'), m.mold_no LIMIT %s OFFSET %s"
        c.execute(sql, params + [page_size, offset])
        rows = c.fetchall()
        return jsonify({'code': 200, 'data': {'list': _fmt_rows(rows), 'total': total, 'page': page, 'pageSize': page_size}})
    finally: conn.close()

# ─── Update Count ───
@molds_bp.route('/api/molds/<int:mold_id>/count', methods=['PUT'])
@login_required
def update_count(mold_id):
    data = request.get_json() or {}
    current_count = data.get('current_count')
    err = validate_positive_int(current_count, '当前模次')
    if err: return jsonify({'code': 400, 'message': '输入有误', 'errors': {'current_count': err}}), 400

    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT id FROM mold_status WHERE mold_id = %s", (mold_id,))
        if not c.fetchone(): return jsonify({'code': 404, 'message': '模具不存在'}), 404
        c.execute("UPDATE mold_status SET current_count = %s WHERE mold_id = %s", (int(current_count), mold_id))
        conn.commit()
        return jsonify({'code': 200, 'message': '模次更新成功'})
    except Exception as e: conn.rollback(); return jsonify({'code': 500, 'message': str(e)}), 500
    finally: conn.close()

# ─── Reset Counter ───
@molds_bp.route('/api/molds/<int:mold_id>/reset-counter', methods=['POST'])
@login_required
def reset_counter(mold_id):
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT * FROM mold_status WHERE mold_id = %s", (mold_id,))
        status = c.fetchone()
        if not status: return jsonify({'code': 404, 'message': '模具不存在'}), 404

        current = status['current_count']
        previous_total = status['total_production']
        new_total = previous_total + current

        # Log the reset
        c.execute("INSERT INTO counter_reset_logs (mold_id, reset_count, previous_total, new_total, reset_date) VALUES (%s, %s, %s, %s, CURDATE())",
                  (mold_id, current, previous_total, new_total))

        # Update status: accumulate to total_production, clear current and last_maintain
        c.execute("""UPDATE mold_status SET total_production = %s, current_count = 0,
            last_maintain_count = 0, last_maintain_date = NULL WHERE mold_id = %s""", (new_total, mold_id))
        conn.commit()
        return jsonify({'code': 200, 'message': '计数器已重置', 'data': {
            'previous_count': current, 'previous_total': previous_total, 'new_total': new_total
        }})
    except Exception as e: conn.rollback(); return jsonify({'code': 500, 'message': str(e)}), 500
    finally: conn.close()

# ─── Maintain ───
@molds_bp.route('/api/molds/maintain', methods=['POST'])
@login_required
def batch_maintain():
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids or not isinstance(ids, list): return jsonify({'code': 400, 'message': '请选择要保养的模具'}), 400
    valid_ids = [int(i) for i in ids if str(i).isdigit()]
    if not valid_ids: return jsonify({'code': 400, 'message': '无效的模具ID'}), 400

    conn = get_db_connection()
    try:
        c = conn.cursor()
        # Get current counts for logging
        placeholders = ','.join(['%s'] * len(valid_ids))
        c.execute(f"SELECT mold_id, current_count, last_maintain_count FROM mold_status WHERE mold_id IN ({placeholders})", valid_ids)
        rows = c.fetchall()

        for row in rows:
            # Log maintenance
            c.execute("""INSERT INTO maintenance_logs (mold_id, previous_maintain_count, new_maintain_count, maintain_date)
                VALUES (%s, %s, %s, CURDATE())""", (row['mold_id'], row['last_maintain_count'], row['current_count']))

        # Update status
        c.execute(f"""UPDATE mold_status SET last_maintain_count = current_count,
            last_maintain_date = CURDATE() WHERE mold_id IN ({placeholders})""", valid_ids)
        conn.commit()
        return jsonify({'code': 200, 'message': f'成功保养 {len(valid_ids)} 个模具'})
    except Exception as e: conn.rollback(); return jsonify({'code': 500, 'message': str(e)}), 500
    finally: conn.close()

# ─── Dashboard ───
@molds_bp.route('/api/molds/dashboard', methods=['GET'])
@login_required
def get_dashboard():
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT COUNT(*) as c FROM molds")
        total = c.fetchone()['c']

        c.execute("""SELECT COUNT(*) as c FROM mold_status ms
            JOIN molds m ON m.id = ms.mold_id
            WHERE (ms.current_count - ms.last_maintain_count) > m.maintain_interval""")
        needs = c.fetchone()['c']

        c.execute("""SELECT COUNT(*) as c FROM mold_status ms
            JOIN molds m ON m.id = ms.mold_id
            WHERE (ms.current_count - ms.last_maintain_count) >= m.maintain_interval * 0.9
            AND (ms.current_count - ms.last_maintain_count) <= m.maintain_interval""")
        warning = c.fetchone()['c']

        normal = total - needs - warning

        # Focused molds as top priority
        sql = f"{STATUS_SQL} WHERE m.is_focused = 1 ORDER BY FIELD(status, 'needs_maintenance', 'warning', 'normal'), m.mold_no LIMIT 10"
        c.execute(sql)
        focused = c.fetchall()

        # If no focused molds, get top by status priority
        if not focused:
            sql = f"{STATUS_SQL} ORDER BY FIELD(status, 'needs_maintenance', 'warning', 'normal') LIMIT 10"
            c.execute(sql)
            focused = c.fetchall()

        return jsonify({'code': 200, 'data': {
            'total': total, 'needs_maintenance': needs, 'warning': warning, 'normal': max(0, normal),
            'top_molds': _fmt_rows(focused)
        }})
    finally: conn.close()

# ─── Counter Reset History ───
@molds_bp.route('/api/molds/<int:mold_id>/reset-history', methods=['GET'])
@login_required
def get_reset_history(mold_id):
    conn = get_db_connection()
    try:
        c = conn.cursor()
        c.execute("SELECT * FROM counter_reset_logs WHERE mold_id = %s ORDER BY reset_date DESC", (mold_id,))
        return jsonify({'code': 200, 'data': _fmt_rows(c.fetchall())})
    finally: conn.close()

# ─── Excel Import ───
@molds_bp.route('/api/molds/import', methods=['POST'])
@login_required
def import_molds():
    print(f"[IMPORT] Request received, files={list(request.files.keys())}, content-type={request.content_type}")
    import sys
    sys.stdout.flush()
    if 'file' not in request.files:
        print(f"[IMPORT] No file in request.files")
        sys.stdout.flush()
        return jsonify({'code': 400, 'message': '请上传Excel文件'}), 400
    file = request.files['file']
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'message': '请上传.xlsx或.xls文件'}), 400

    try:
        # Read file content into memory first to avoid FileStorage stream issues
        file_content = file.read()
        print(f"[IMPORT] filename={file.filename}, size={len(file_content) if file_content else 0} bytes")
        if file_content and len(file_content) > 4:
            print(f"[IMPORT] file header bytes: {file_content[:8].hex()}")
        sys.stdout.flush()
        if not file_content or len(file_content) < 100:
            return jsonify({'code': 400, 'message': '上传文件内容为空或无效'}), 400
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        required = ['模具编号', '模具名称', '保养模次间隔']
        for r in required:
            if r not in headers: return jsonify({'code': 400, 'message': f'模板缺少必要列：{r}'}), 400

        idx_map = {h: i for i, h in enumerate(headers)}
        results = {'success': 0, 'failed': 0, 'errors': []}

        conn = get_db_connection()
        try:
            c = conn.cursor()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row): continue
                mold_no = str(row[idx_map.get('模具编号', 0)] or '').strip().upper()
                mold_name = str(row[idx_map.get('模具名称', 1)] or '').strip()
                interval = row[idx_map.get('保养模次间隔', 2)]
                current = row[idx_map.get('当前模次', 3)] or 0
                last_maintain = row[idx_map.get('上次保养模次', 4)] or 0
                last_date = row[idx_map.get('上次保养日期', 5)] or None
                total_prod = row[idx_map.get('累计生产模次', 6)] or 0

                if not mold_no or not mold_name: continue
                try: interval = int(interval)
                except: results['errors'].append(f'{mold_no}: 保养间隔无效'); results['failed'] += 1; continue

                c.execute("SELECT id FROM molds WHERE mold_no = %s", (mold_no,))
                if c.fetchone(): results['errors'].append(f'{mold_no}: 编号已存在'); results['failed'] += 1; continue

                c.execute("INSERT INTO molds (mold_no, mold_name, maintain_interval) VALUES (%s, %s, %s)",
                          (mold_no, mold_name, interval))
                new_id = c.lastrowid
                c.execute("""INSERT INTO mold_status (mold_id, current_count, last_maintain_count,
                    last_maintain_date, total_production) VALUES (%s, %s, %s, %s, %s)""",
                    (new_id, int(current or 0), int(last_maintain or 0),
                     last_date if last_date else None, int(total_prod or 0)))
                results['success'] += 1
            conn.commit()
        except Exception as e: conn.rollback(); raise
        finally: conn.close()

        return jsonify({'code': 200, 'message': f'导入完成：成功 {results["success"]} 条，失败 {results["failed"]} 条', 'data': results})
    except Exception as e:
        err_msg = str(e)
        # Diagnostic: show what was actually received
        diag_parts = [f"filename={file.filename}"]
        if 'file_content' in dir() and file_content:
            diag_parts.append(f"size={len(file_content)} bytes")
            diag_parts.append(f"header_hex={file_content[:32].hex()}")
            # Try to decode as text for common formats
            try:
                text_preview = file_content[:200].decode('utf-8', errors='replace')
                if text_preview.strip():
                    diag_parts.append(f"text_preview={text_preview[:100]}")
            except:
                pass
        else:
            diag_parts.append("file_content=N/A")
        diag = " | ".join(diag_parts)
        return jsonify({'code': 500, 'message': f'导入失败: {err_msg} | 诊断: {diag}'}), 500

@molds_bp.route('/api/molds/import-template', methods=['GET'])
@login_required
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '模具导入模板'
    headers = ['模具编号', '模具名称', '保养模次间隔', '当前模次', '上次保养模次', '上次保养日期', '累计生产模次']
    ws.append(headers)

    # Header style
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Example data
    ws.append(['M005', '上壳模具E', '6000', '12000', '8000', '2025-01-15', '50000'])
    ws.append(['M006', '下壳模具F', '4000', '8000', '5000', '2025-02-20', '30000'])

    # Column widths
    for i, w in enumerate([14, 16, 16, 12, 16, 16, 16], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue(), 200, {
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'Content-Disposition': 'attachment; filename=模具导入模板.xlsx'
    }
